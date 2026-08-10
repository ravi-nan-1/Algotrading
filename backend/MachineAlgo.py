import ssl
import certifi
from py5paisa import FivePaisaClient
from py5paisa.order import Order, OrderType, Exchange
import pyotp
import os
import mibian as mb
working_dir = os.path.dirname(os.path.abspath(__file__))
import pandas as pd
import datetime as dt
import auth
import pandas_ta as ta
import numpy as np
import requests
import Telegram_token
import math
from openpyxl import load_workbook
import json
import threading
import pytz
import joblib
from sklearn.ensemble import RandomForestClassifier, RandomForestRegressor
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report
import joblib
UTC = pytz.timezone('Asia/Kolkata')
import time
from pymongo import MongoClient
import sklearn
import requests

from collections import deque, Counter

try:
    import pandas_ta as ta
except ImportError as e:
    raise ImportError(
        "pandas_ta is required for the multi-indicator divergence engine "
        "(pip install pandas_ta)"
    ) from e

try:
    from scipy.signal import argrelextrema
except ImportError as e:
    raise ImportError(
        "scipy is required for the standalone find_swing_levels() utility "
        "(pip install scipy)"
    ) from e


ssl_context = ssl.create_default_context(cafile=certifi.where())

print("scikit-learn version:", sklearn.__version__)
uri = "mongodb+srv://singhrajeev1470_db_user:kaPh8sxuaVFWWsSr@cluster0.mtmtbrr.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0&tls=true"

# Connect to MongoDB
client1 = MongoClient(uri, tls=True, tlsCAFile=certifi.where())

# Choose a database (it will be created if not exist)
db = client1['AlgoTradingDB']
print(db.list_collection_names())
ce_collection = db["CE_Options"]
pe_collection = db["PE_Options"]

client = FivePaisaClient(cred=auth.cred)
print(pyotp.TOTP(auth.token).now())

# New TOTP based authentication
client.get_totp_session(auth.client_id, pyotp.TOTP(auth.token).now(), auth.pin)

# User Inputs
START_TIME = [9, 16, 0]  # Algo Start Time
EXIT_TIME = [23, 30, 0]  # Algo End Time

Total_Cash = 10000
Max_Position = 1
Total_Cash_per_position = int(Total_Cash / Max_Position)

Take_Profit = 20

#Tickers = ['NIFTY 08 MAY 2025 CE 24400.00','NIFTY 08 MAY 2025 PE 24400.00']


import json
import os

ticker_path = os.path.join(os.path.dirname(__file__), "tickers.json")
print("🔍 Looking for:", ticker_path)

if os.path.exists(ticker_path):
    print("📂 Loading tickers.json...")
    with open(ticker_path, "r") as f:
        Tickers = json.load(f)


# Getting Instrument
instrument_df = pd.read_csv('ScripMasterfno.csv')
instrument_df = instrument_df[(instrument_df.Exch == 'N')]
#print(instrument_df[(instrument_df.Name == 'NIFTY 03 APR 2025 CE 23300.00')])








signal_data = []

# Getting Script Code
def scripcode_lookup(instrument=instrument_df, symbol='TCS'):
    ## This function is used to find the instrument token number
    try:
        return instrument[instrument.Name == symbol].ScripCode.values[0]
    except:
        return -1



def get_cash_market_data_3(timeframe='3m'):
    df = pd.DataFrame(client.historical_data(Exch='N', ExchangeSegment='C', ScripCode=999920000, time=timeframe,
                                             From=dt.date.today() - dt.timedelta(5), To=dt.date.today()))

    print(df.columns)
    df.set_index("Datetime", inplace=True)
    print(df)
    return df



def get_cash_market_data(symbol, timeframe):
    scriptcode = scripcode_lookup(instrument_df, symbol)
    sym=symbol

    parts = symbol.split()
    ticker = parts[0]  # Extract ticker
    expiry = f"{parts[1]} {parts[2]} {parts[3]}"  # Extract expiry date
    opttype = parts[4]  # Extract option type (CE/PE)
    strike = float(parts[5])  # Extract strike price and convert to float

    df = pd.DataFrame(client.historical_data(Exch='N', ExchangeSegment='D', ScripCode=scriptcode, time=timeframe,
                                             From=dt.date.today()-dt.timedelta(3), To=dt.date.today()))

    df.set_index("Datetime", inplace=True)
    df["Option_Type"] = opttype
    df["Strike_Price"] = strike
    



    print(df)
    return df







def opt_exp(ticker):
    # Filter the relevant data for the given ticker and options
    dates = instrument_df[
        (instrument_df.SymbolRoot == ticker) & ((instrument_df.ScripType == 'CE') | (instrument_df.ScripType == 'PE'))]

    # Get the unique expiry dates and convert them to datetime objects
    dates = dates['Expiry'].unique().tolist()

    dates = [dt.datetime.strptime(date, '%Y-%m-%d') for date in dates]

    # Get today's date
    today = dt.datetime.today()

    # Sort the dates in ascending order
    dates.sort()

    # Find the next available date after today (skip today's date)
    future_dates = [date for date in dates if date > today]

    if future_dates:
        # Get the next available future date
        trade = future_dates[0]
    else:
        # If no future date is found, return None or handle the fallback
        return "No future expiry dates available."

    # Return the selected expiry date in the desired format
    return trade.strftime('%d %b %Y')



def process_expiry_date(date_str):
    # Extract the timestamp from the '/Date(...)' format
    timestamp = int(date_str.split('(')[1].split('+')[0])

    # Convert the timestamp to a datetime object
    date = dt.datetime.utcfromtimestamp(timestamp / 1000.0)

    # Convert datetime to the required format
    formatted_date = date.strftime('%d %b %Y')

    return timestamp, formatted_date


DELTA = 30

def fetch_option_data(option_string):
    print(option_string)
    expiry1 = opt_exp("NIFTY")
    print(expiry1)
    parts = option_string.split()
    ticker = parts[0]  # Extract ticker
    expiry = f"{parts[1]} {parts[2]} {parts[3]}"  # Extract expiry date
    opttype = parts[4]  # Extract option type (CE/PE)
    strike = float(parts[5])  # Extract strike price and convert to float

    target_strike = int(strike)  # Convert strike price to integer

    a = client.get_expiry("N", ticker)
    expiry_list = pd.DataFrame(a['Expiry'])
    spot_price = a['lastrate'][0]['LTP']

    # Process expiry dates
    expiry_list['Timestamp'], expiry_list['Format'] = zip(*expiry_list['ExpiryDate'].apply(process_expiry_date))

    # Get timestamp for target expiry
    timestamp_row = expiry_list[expiry_list.Format == expiry1]
    if timestamp_row.empty:
        print("Error: Expiry date not found")
        return None
    timestamp = timestamp_row.Timestamp.values[0]

    # Fetch option chain for the expiry
    option_chain = client.get_option_chain("N", ticker, timestamp)
    option_chain = pd.DataFrame(option_chain['Options'])

    # Filter for specified option type (CE or PE) & non-zero last traded price
    option_chain = option_chain[(option_chain.CPType == opttype) & (option_chain.LastRate != 0)]

    # Filter only for the target strike price
    option_chain = option_chain[option_chain.StrikeRate == target_strike]

    if option_chain.empty:
        print("Error: No data for target strike price")
        return None

    option_chain['SPOT'] = spot_price
    startTime = dt.datetime.today()
    date_obj = dt.datetime.strptime(expiry, "%d %b %Y")
    daysToExpiry = max((date_obj-startTime).days, 1)  # Ensure non-negative days

    # Create DataFrame
    opt_data = pd.DataFrame()
    opt_data['SPOT'] = option_chain['SPOT']
    opt_data['STRIKE'] = option_chain['StrikeRate']
    opt_data[f'{opttype}_LTP'] = option_chain['LastRate']
    opt_data['OI'] = option_chain['OpenInterest']
    opt_data['SYMBOL'] = option_chain['Name']
    opt_data = opt_data.reset_index(drop=True)

    Delta, Gamma, Theta, IV = [], [], [], []

    # Calculate Implied Volatility, Delta, Gamma, Theta
    r = 10  # Risk-free rate
    for i in range(len(opt_data)):
        c = mb.BS([opt_data['SPOT'][i], opt_data['STRIKE'][i], r, daysToExpiry],
                  callPrice=opt_data[f'{opttype}_LTP'][i])
        civ = c.impliedVolatility  # Fetch implied volatility
        cg = mb.BS([opt_data['SPOT'][i], opt_data['STRIKE'][i], r, daysToExpiry], volatility=civ)

        if opttype == 'CE':
            Delta.append(cg.callDelta * 100)
            Theta.append(cg.callTheta)
        else:
            Delta.append(cg.putDelta * 100)
            Theta.append(cg.putTheta)

        Gamma.append(cg.gamma * 100)  # Convert to percentage
        IV.append(civ)  # Store IV

    # Storing calculated Greeks in DataFrame
    opt_data[f'{opttype}_Delta'] = Delta
    opt_data[f'{opttype}_Gamma'] = Gamma
    opt_data[f'{opttype}_Theta'] = Theta
    opt_data['Implied_Volatility'] = IV

    opt_data["inserted_at"] = dt.datetime.now(UTC).strftime("%d-%b-%Y %I:%M%p")
    records = opt_data.to_dict("records")

    # Save to Excel
    if opttype == 'CE':
        db['NIFTY_CE'].insert_many(opt_data.to_dict('records'))
    else:
        db['NIFTY_PE'].insert_many(opt_data.to_dict('records'))

    return opt_data



def volume_oscillator(df, fast=14, slow=28):
    ema_fast = df['Volume'].ewm(span=fast, adjust=False).mean()
    ema_slow = df['Volume'].ewm(span=slow, adjust=False).mean()
    vo = ((ema_fast - ema_slow) / ema_slow) * 100
    return vo


def detect_trendline_touches_for_strategy(df: pd.DataFrame,
                                          swing_period: int = 8,
                                          min_hl_points: int = 2,
                                          lookback_candles: int = 50,
                                          tolerance: float = 0.007,
                                          extend_back: bool = True):
    """
    Trendline touch detector that extends line BACKWARDS to find ALL touches

    Args:
        df: DataFrame with OHLC data
        swing_period: Candles to look left/right (8 = 24 minutes for 3m)
        min_hl_points: Minimum Higher Lows to connect (2-3)
        lookback_candles: Analyze last N candles (50 = 2.5 hours for 3m)
        tolerance: Touch tolerance (0.007 = 0.7%)
        extend_back: Extend trendline backwards to find earlier touches

    Returns:
        List of indices where touches occur
    """

    # Only analyze recent candles
    if len(df) > lookback_candles:
        start_idx = len(df)-lookback_candles
        df_recent = df.iloc[start_idx:].copy()
        offset = start_idx
    else:
        df_recent = df.copy()
        offset = 0

    # Find swing lows
    swing_lows = []
    for i in range(swing_period, len(df_recent)):
        # Allow recent candles without full right-side confirmation
        if i+swing_period > len(df_recent):
            current_low = df_recent['Low'].iloc[i]
            left_lows = df_recent['Low'].iloc[max(0, i-swing_period):i]

            if len(left_lows) > 0 and current_low <= left_lows.min():
                swing_lows.append((i+offset, current_low))
        else:
            current_low = df_recent['Low'].iloc[i]
            left_lows = df_recent['Low'].iloc[i-swing_period:i]
            right_lows = df_recent['Low'].iloc[i+1:min(len(df_recent), i+swing_period+1)]

            if current_low <= left_lows.min() and current_low <= right_lows.min():
                swing_lows.append((i+offset, current_low))

    if len(swing_lows) < 2:
        # print(f"⚠️ Only found {len(swing_lows)} swing lows. Need at least 2.")
        return []

    # print(f"✓ Found {len(swing_lows)} swing lows")

    # Filter for Higher Lows
    higher_lows = [swing_lows[0]]
    for i in range(1, len(swing_lows)):
        idx, low = swing_lows[i]
        prev_idx, prev_low = higher_lows[-1]

        # Accept if higher or nearly equal (0.2% tolerance)
        if low >= prev_low * 0.998:
            higher_lows.append((idx, low))

    if len(higher_lows) < min_hl_points:
        # print(f"⚠️ Only found {len(higher_lows)} Higher Lows. Need at least {min_hl_points}.")
        return []

    # print(f"✓ Found {len(higher_lows)} Higher Lows")
    # print(f"   Points: {[(idx, round(price, 2)) for idx, price in higher_lows]}")

    # Take most recent Higher Lows for trendline
    recent_hl = higher_lows[-min_hl_points:]
    idx1, price1 = recent_hl[0]
    idx2, price2 = recent_hl[-1]

    # Calculate trendline
    slope = (price2-price1) / (idx2-idx1)

    # Accept flat or upward lines
    if slope < -0.0001:
        # print(f"⚠️ Negative slope: {slope:.6f}. Not an uptrend.")
        return []

    intercept = price1-slope * idx1


    print(f"✓ Trendline: slope={slope:.6f}, intercept={intercept:.2f}")
    print(f"   Line connects indices {idx1} to {idx2}")

    # Extend trendline backwards
    if extend_back:
        search_start = max(0, len(df)-lookback_candles)
        # print(f"   Extending line backwards from index {idx1} to index {search_start}")
    else:
        search_start = idx1

    # Detect touches from the START of lookback period to END of data
    touch_indices = []

    for i in range(search_start, len(df)):
        line_value = slope * i+intercept
        threshold = line_value * tolerance

        current_low = df['Low'].iloc[i]
        current_high = df['High'].iloc[i]
        current_close = df['Close'].iloc[i]

        # Skip anchor points
        is_anchor_point = any(i == hl_idx for hl_idx, _ in recent_hl)

        if i > 0 and not is_anchor_point:
            prev_close = df['Close'].iloc[i-1]

            # Detection methods
            touch_from_above = (prev_close > line_value and
                                current_low <= line_value+threshold)

            crosses_line = (current_low <= line_value+threshold and
                            current_high >= line_value-threshold)

            close_near_line = abs(current_close-line_value) <= threshold

            low_touches = abs(current_low-line_value) <= threshold

            # Detect touch
            if touch_from_above or crosses_line or close_near_line or low_touches:
                # Additional validation: don't add if price breaks significantly below
                if current_close > line_value-(threshold * 2):
                    touch_indices.append(i)
                    # touch_type = 'from_above' if touch_from_above else ('cross' if crosses_line else ('close' if close_near_line else 'low_bounce'))

                    # Only print recent touches to avoid spam
                    # if i >= len(df) - 20:
                    #     print(f"   Touch at index {i}: Low={current_low:.2f}, Close={current_close:.2f}, Line={line_value:.2f} ({touch_type})")

    # print(f"✓ Found {len(touch_indices)} total touches (from index {search_start} to {len(df)-1})")

    # Show distribution of touches
    # if len(touch_indices) > 0:
    #     print(f"   First touch at index: {touch_indices[0]}")
    #     print(f"   Last touch at index: {touch_indices[-1]}")
    #     print(f"   Touch indices: {touch_indices}")

    return touch_indices










GROQ_API_KEY = os.getenv("GROQ_API_KEY")










def fetch_ohlcv(symbol, timeframe="3m", candles=30, existing_data=None):
    """
    Fetch OHLCV data for LLM analysis.
    If existing_data is provided, use it directly.
    Otherwise fetch from 5paisa API.
    """
    print('📡 fetch_ohlcv called')

    # ═══════════════════════════════════════════════════════════════
    #  PATH 1: Use existing data (same data super_trend analyzed)
    # ═══════════════════════════════════════════════════════════════
    if existing_data is not None and isinstance(existing_data, pd.DataFrame):
        print('  ✅ Using existing DataFrame (same as super_trend input)')

        required = {"Open", "High", "Low", "Close", "Volume"}
        if not required.issubset(existing_data.columns):
            print(f"  ❌ Missing columns: {required - set(existing_data.columns)}")
            return None

        df = existing_data.tail(candles).copy()

        ohlcv = []
        for _, r in df.iterrows():
            try:
                ohlcv.append({
                    "o": round(float(r["Open"]), 2),
                    "h": round(float(r["High"]), 2),
                    "l": round(float(r["Low"]), 2),
                    "c": round(float(r["Close"]), 2),
                    "v": int(float(r["Volume"]))
                })
            except (ValueError, TypeError) as e:
                print(f"  ⚠️ Skipping bad row: {e}")
                continue

        if len(ohlcv) == 0:
            print("  ❌ No valid OHLCV rows")
            return None

        print(f"  📊 Converted {len(ohlcv)} candles from existing data")
        return ohlcv

    # ═══════════════════════════════════════════════════════════════
    #  PATH 2: Fresh API fetch (fallback)
    # ═══════════════════════════════════════════════════════════════
    print('  📡 Fetching fresh from 5paisa API...')

    scripcode = scripcode_lookup(instrument_df, symbol)
    print(f'  ScripCode: {scripcode}')
    if not scripcode:
        print("  ❌ ScripCode not found")
        return None

    try:
        raw = client.historical_data(
            Exch='N', ExchangeSegment='D',
            ScripCode=scripcode,
            time=timeframe,
            From=dt.date.today() - dt.timedelta(5),
            To=dt.date.today()
        )
    except Exception as e:
        print(f"  ❌ API call failed: {e}")
        return None

    # ═══════════════════════════════════════════════════════════════
    #  🔥 FIX: Handle both DataFrame and list/dict returns
    # ═══════════════════════════════════════════════════════════════
    if isinstance(raw, pd.DataFrame):
        df = raw.copy()
    elif isinstance(raw, list):
        if len(raw) == 0:
            print("  ❌ API returned empty list")
            return None
        df = pd.DataFrame(raw)
    elif isinstance(raw, dict):
        df = pd.DataFrame([raw])
    else:
        print(f"  ❌ Unexpected API return type: {type(raw)}")
        return None

    # Check if DataFrame is empty
    if df.empty:
        print("  ❌ API returned empty DataFrame")
        return None

    print(f"  📦 API returned {len(df)} rows")
    print(f"  📋 Columns: {list(df.columns)}")

    # ─── Datetime handling ───
    datetime_col = None
    for col in ["Datetime", "DateTime", "Date", "Time", "datetime", "date"]:
        if col in df.columns:
            datetime_col = col
            break

    if datetime_col:
        try:
            df[datetime_col] = pd.to_datetime(df[datetime_col])
            df = df.sort_values(datetime_col)
        except Exception as e:
            print(f"  ⚠️ DateTime parse warning: {e}")

    # ─── Column normalization ───
    col_map = {}
    for col in df.columns:
        cl = col.lower().strip()
        if cl in ('open', 'o'):
            col_map[col] = 'Open'
        elif cl in ('high', 'h'):
            col_map[col] = 'High'
        elif cl in ('low', 'l'):
            col_map[col] = 'Low'
        elif cl in ('close', 'c'):
            col_map[col] = 'Close'
        elif cl in ('volume', 'v', 'vol'):
            col_map[col] = 'Volume'

    if col_map:
        df = df.rename(columns=col_map)

    # ─── Validate columns ───
    required = {"Open", "High", "Low", "Close", "Volume"}
    if not required.issubset(df.columns):
        missing = required - set(df.columns)
        print(f"  ❌ Missing columns: {missing}")
        print(f"  Available: {list(df.columns)}")
        return None

    # ─── Drop extra columns, keep only OHLCV ───
    df = df[['Open', 'High', 'Low', 'Close', 'Volume']].copy()
    df = df.tail(candles)

    # ─── Convert to list of dicts ───
    ohlcv = []
    for _, r in df.iterrows():
        try:
            ohlcv.append({
                "o": round(float(r["Open"]), 2),
                "h": round(float(r["High"]), 2),
                "l": round(float(r["Low"]), 2),
                "c": round(float(r["Close"]), 2),
                "v": int(float(r["Volume"]))
            })
        except (ValueError, TypeError) as e:
            print(f"  ⚠️ Skipping bad row: {e}")
            continue

    if len(ohlcv) == 0:
        print("  ❌ No valid OHLCV rows")
        return None

    print(f"  ✅ Returning {len(ohlcv)} candles")
    return ohlcv





def llm_trade_signal(symbol, ohlcv, timeframe, signal_context=None):
    """
    ADVANCED LLM FILTER — REJECT BAD TRADES ONLY

    Role:
    - DO NOT generate signals
    - DO NOT overthink indicators
    - ONLY detect traps & risky setups

    Output:
    - "buy" → allow trade
    - "no trade" → reject trade
    """

    print('🧠 llm_trade_signal (ADVANCED FILTER MODE)')

    if not ohlcv or len(ohlcv) < 12:
        return _no_trade_response("Not enough candles")

    closes  = [c["c"] for c in ohlcv]
    highs   = [c["h"] for c in ohlcv]
    lows    = [c["l"] for c in ohlcv]
    opens   = [c["o"] for c in ohlcv]
    volumes = [c["v"] for c in ohlcv]

    current_price = closes[-1]

    # ─────────────────────────────────────────────
    # 🔹 Compact Structured Context (IMPORTANT)
    # ─────────────────────────────────────────────
    last_6 = [
        {
            "o": round(opens[i],2),
            "h": round(highs[i],2),
            "l": round(lows[i],2),
            "c": round(closes[i],2)
        }
        for i in range(-6, 0)
    ]

    avg_vol_5 = sum(volumes[-5:]) / 5
    vol_ratio = round(volumes[-1] / avg_vol_5, 2) if avg_vol_5 > 0 else 1

    recent_high = max(highs[-10:])
    recent_low  = min(lows[-10:])
    range_pos = round((current_price - recent_low) / (recent_high - recent_low), 2) if recent_high != recent_low else 0.5

    primary_reason = signal_context.get("reason","") if signal_context else ""

    # ─────────────────────────────────────────────
    # 🔥 ADVANCED PROMPT (THIS IS THE EDGE)
    # ─────────────────────────────────────────────
    prompt = f"""
You are an ELITE OPTIONS SCALPING RISK MANAGER.

You DO NOT give buy signals.
You ONLY REJECT bad trades.

Analyze the structure deeply like a professional trader.

DATA:
Candles (last 6):
{last_6}

Volume ratio (current / avg5): {vol_ratio}
Range position (0=low, 1=high): {range_pos}
Recent High: {recent_high}
Recent Low: {recent_low}

Primary Signal Reason:
{primary_reason}

----------------------------------------

REJECT TRADE if you see ANY of these:

1. CHOPPY MARKET:
- overlapping candles
- no clear direction
- alternating colors

2. FAKE BREAKOUT:
- price near high but weak close
- upper wicks / rejection
- breakout without volume follow-through

3. EXHAUSTION MOVE:
- 2-3 strong candles already done
- late entry near top

4. LOW QUALITY MOMENTUM:
- small bodies
- inconsistent candle structure

5. BAD LOCATION:
- range_pos > 0.75 (too close to resistance)
- no room to move

----------------------------------------

ALLOW TRADE ONLY IF:
- clean directional candles
- strong closes near highs
- no rejection wicks
- not extended already

----------------------------------------

STRICT RULE:
When in doubt → REJECT TRADE

----------------------------------------

Return ONLY JSON:

{{
  "signal": "buy" or "no trade",
  "confidence": 0.0-1.0,
  "reason": "very short explanation"
}}
"""

    try:
        response = requests.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {GROQ_API_KEY}",
                "Content-Type": "application/json"
            },
            json={
                "model": "llama-3.3-70b-versatile",
                "temperature": 0.0,   # ❗ ultra deterministic
                "max_tokens": 120,
                "messages": [
                    {"role": "system", "content": (
                        "You are a strict risk manager. "
                        "You reject bad trades aggressively. "
                        "You NEVER hallucinate. "
                        "Output ONLY valid JSON."
                    )},
                    {"role": "user", "content": prompt}
                ]
            },
            timeout=10
        )

        if response.status_code != 200:
            return _no_trade_response(f"API error {response.status_code}")

        content = response.json()["choices"][0]["message"]["content"].strip()

        if "```" in content:
            content = content.replace("```json","").replace("```","")

        result = json.loads(content[content.index("{"):content.rindex("}")+1])

        signal = str(result.get("signal","no trade")).lower()
        confidence = float(result.get("confidence",0.0))

        if signal not in ("buy","no trade"):
            signal = "no trade"

        # ─────────────────────────────────────────────
        # 🔒 FINAL SAFETY (VERY IMPORTANT)
        # ─────────────────────────────────────────────
        if confidence < 0.55:
            signal = "no trade"

        entry = current_price
        stop_loss = round(current_price - 10, 2)
        target_1  = round(current_price + 20, 2)
        target_2  = round(current_price + 30, 2)

        final = {
            "signal": signal,
            "confidence": round(confidence,2),
            "entry": entry,
            "stop_loss": stop_loss,
            "target_1": target_1,
            "target_2": target_2,
            "reason": result.get("reason","")
        }

        print(f"  🧠 FILTER RESULT: {signal} | {confidence} | {final['reason']}")
        return final

    except Exception as e:
        return _no_trade_response(f"LLM error: {e}")

def super_trend(
    symbol: str,
    data: pd.DataFrame,
    pivot_period: int = 5,
    min_count: int = 1,
    maxpp: int = 10,
    maxbars: int = 100,
    search: str = "Regular",              # "Regular" | "Hidden" | "Regular/Hidden"
    showlast: bool = False,
    dontconfirm: bool = False,
    source: str = "Close",                # "Close" (Pine default) or anything else -> Low
    start_time: dt.time = dt.time(10, 0),
    stoch_oversold: float = 30.0,
    min_pivot_atr_mult: float = 0.0,
    cooldown_candles: int = 1,
    verbose: bool = True,
) -> pd.DataFrame:


    # ------------------------------------------------------------------
    # Nested helpers
    # ------------------------------------------------------------------
    def _calculate_indicators(df):
        close = df['Close']
        high = df['High']
        low = df['Low']
        volume = df['Volume']

        df['rsi'] = ta.rsi(close, length=14)

        macd_df = ta.macd(close, fast=12, slow=26, signal=9)
        df['macd'] = macd_df.iloc[:, 0]
        df['deltamacd'] = macd_df.iloc[:, 1]

        df['moment'] = close.diff(10)
        df['cci'] = ta.cci(high, low, close, length=10)
        df['obv'] = ta.obv(close, volume)

        stoch_df = ta.stoch(high, low, close, k=14, d=3, smooth_k=1)
        df['stk'] = stoch_df.iloc[:, 0].rolling(3).mean()

        vwma_fast = (close * volume).rolling(12).sum() / volume.rolling(12).sum()
        vwma_slow = (close * volume).rolling(26).sum() / volume.rolling(26).sum()
        df['vwmacd'] = vwma_fast - vwma_slow

        hl = (high - low).replace(0, np.nan)
        cmfm = ((close - low) - (high - close)) / hl
        cmfv = cmfm * volume
        df['cmf'] = cmfv.rolling(21).sum() / volume.rolling(21).sum()

        df['mfi'] = ta.mfi(high, low, close, volume, length=14)

        return df

    def _pine_pivotlow(arr, i, prd):
        if i < 2 * prd:
            return None
        center = i - prd
        window = arr[i - 2 * prd: i + 1]
        valid_window = window[~np.isnan(window)]
        if len(valid_window) == 0:
            return None
        center_val = arr[center]
        if np.isnan(center_val):
            return None
        if center_val == np.min(valid_window):
            return center_val
        return None

    def _nz(val, default=0.0):
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return default
        return val

    def _get_back(arr, i, n):
        idx = i - n
        if idx < 0 or idx >= len(arr):
            return np.nan
        return arr[idx]

    def _positive_divergence(i, src, close, prsc, pl_positions, pl_vals,
                              prd, maxpp, maxbars, dontconfirm, cond):
        """cond=1 -> regular positive (bullish) divergence.
        cond=2 -> hidden positive (bullish) divergence."""
        divlen = 0

        src_0 = _get_back(src, i, 0)
        src_1 = _get_back(src, i, 1)
        close_0 = _get_back(close, i, 0)
        close_1 = _get_back(close, i, 1)

        gate = (
            dontconfirm or
            (not np.isnan(src_0) and not np.isnan(src_1) and src_0 > src_1) or
            (not np.isnan(close_0) and not np.isnan(close_1) and close_0 > close_1)
        )
        if not gate:
            return divlen

        startpoint = 0 if dontconfirm else 1
        n_pivots = len(pl_positions)

        for x in range(maxpp):
            if x >= n_pivots:
                break
            pl_pos = pl_positions[x]
            pl_val = pl_vals[x]
            if pl_pos == 0:
                break

            length = i - pl_pos + prd
            if length > maxbars:
                break
            if length <= 5:
                continue

            src_sp = _get_back(src, i, startpoint)
            src_len = _get_back(src, i, length)
            prsc_sp = _get_back(prsc, i, startpoint)

            if np.isnan(src_sp) or np.isnan(src_len) or np.isnan(prsc_sp):
                continue

            cond1_check = (cond == 1 and src_sp > src_len and prsc_sp < _nz(pl_val))
            cond2_check = (cond == 2 and src_sp < src_len and prsc_sp > _nz(pl_val))
            if not (cond1_check or cond2_check):
                continue

            close_sp = _get_back(close, i, startpoint)
            close_len = _get_back(close, i, length)
            if np.isnan(close_sp):
                continue

            span = length - startpoint
            if span == 0:
                continue

            slope1 = (src_sp - src_len) / span
            virtual_line1 = src_sp - slope1
            slope2 = (close_sp - _nz(close_len)) / span
            virtual_line2 = close_sp - slope2

            arrived = True
            for y in range(1 + startpoint, length):
                src_y = _get_back(src, i, y)
                close_y = _get_back(close, i, y)
                if np.isnan(src_y):
                    arrived = False
                    break
                close_y_nz = _nz(close_y)
                if src_y < virtual_line1 or close_y_nz < virtual_line2:
                    arrived = False
                    break
                virtual_line1 -= slope1
                virtual_line2 -= slope2

            if arrived:
                divlen = length
                break

        return divlen


    if 'Datetime' in data.columns:
        data = data.set_index('Datetime')
    if not isinstance(data.index, pd.DatetimeIndex):
        data.index = pd.to_datetime(data.index)
    if data.index.tz is not None:
        if verbose:
            print(f"⚠️  Datetime index is tz-aware ({data.index.tz}) -- stripping tz.")
        data.index = data.index.tz_localize(None)
    data.index.name = 'Datetime'

    data = data.copy()
    data = data[~data.index.duplicated(keep='last')].sort_index()
    n = len(data)

    if verbose:
        print("=" * 100)
        print(f"🚀 MULTI-INDICATOR BULLISH DIVERGENCE ENGINE → {symbol} | Data Points: {n}")
        print("=" * 100)

    data['st_sig'] = 0
    data['agree_count'] = 0
    data['divergence_indicators'] = ''
    data['reason'] = ''

    data = _calculate_indicators(data)

    close = data['Close'].values.astype(float)
    low = data['Low'].values.astype(float)

    _hl = data['High'] - data['Low']
    _hc = (data['High'] - data['Close'].shift()).abs()
    _lc = (data['Low'] - data['Close'].shift()).abs()
    atr_arr = pd.concat([_hl, _hc, _lc], axis=1).max(axis=1).rolling(14).mean().values

    def _pivot_is_significant(i, val, opp_positions, opp_vals):
        if min_pivot_atr_mult <= 0:
            return True
        if len(opp_positions) == 0 or opp_positions[0] == 0:
            return True
        atr_now = atr_arr[i] if i < len(atr_arr) else np.nan
        if np.isnan(atr_now) or atr_now == 0:
            return True
        amplitude = abs(val - opp_vals[0])
        return amplitude >= (atr_now * min_pivot_atr_mult)

    pl_source = close if source == "Close" else low
    prsc_pos = close if source == "Close" else low

    ind_list = [
        ("MACD", data['macd'].values.astype(float)),
        ("Hist", data['deltamacd'].values.astype(float)),
        ("RSI", data['rsi'].values.astype(float)),
        ("Stoch", data['stk'].values.astype(float)),
        ("CCI", data['cci'].values.astype(float)),
        ("MOM", data['moment'].values.astype(float)),
        ("OBV", data['obv'].values.astype(float)),
        ("VWMACD", data['vwmacd'].values.astype(float)),
        ("CMF", data['cmf'].values.astype(float)),
        ("MFI", data['mfi'].values.astype(float)),
    ]

    do_regular = search in ("Regular", "Regular/Hidden")
    do_hidden = search in ("Hidden", "Regular/Hidden")

    MAXARR = 20
    pl_positions = deque([0] * MAXARR, maxlen=MAXARR)
    pl_vals = deque([0.0] * MAXARR, maxlen=MAXARR)
    # No pivot-high tracking in this long-only build -- see module docstring.
    _no_opp_positions = deque([0] * MAXARR, maxlen=MAXARR)
    _no_opp_vals = deque([0.0] * MAXARR, maxlen=MAXARR)

    remove_last_pos_divs = False
    pos_label_history = []   # each: {'bar', 'agree_count', 'indicators'}

    for i in range(n):
        pl = _pine_pivotlow(pl_source, i, pivot_period)

        if pl is not None:
            if _pivot_is_significant(i, pl, _no_opp_positions, _no_opp_vals):
                pl_positions.appendleft(i)
                pl_vals.appendleft(pl)
            remove_last_pos_divs = False

        if i < 2 * pivot_period:
            continue

        pos_reg = pos_hid = 0
        active_names = []

        for name, indicator in ind_list:
            reg_hit = hid_hit = 0
            if do_regular:
                reg_hit = _positive_divergence(
                    i, indicator, close, prsc_pos, pl_positions, pl_vals,
                    pivot_period, maxpp, maxbars, dontconfirm, cond=1
                )
            if do_hidden:
                hid_hit = _positive_divergence(
                    i, indicator, close, prsc_pos, pl_positions, pl_vals,
                    pivot_period, maxpp, maxbars, dontconfirm, cond=2
                )
            if reg_hit:
                pos_reg += 1
                active_names.append(name)
            if hid_hit:
                pos_hid += 1
                if name not in active_names:
                    active_names.append(name)

        total = pos_reg + pos_hid
        if total < min_count:
            continue

        if showlast:
            pos_label_history.clear()
        else:
            if remove_last_pos_divs and pos_label_history:
                pos_label_history.pop()

        pos_label_history.append({
            'bar': i,
            'agree_count': total,
            'indicators': ', '.join(active_names),
        })
        remove_last_pos_divs = True

    if verbose:
        print(f"🔎 Bullish divergence labels found: {len(pos_label_history)}")

    # ------------------------------------------------------------------
    # Apply trading gates (time-of-day, Stochastic oversold, cooldown)
    # and write final st_sig = 1 entries.
    # ------------------------------------------------------------------
    last_signal_idx = -(10 ** 9)
    fired = 0

    for label in pos_label_history:
        i = label['bar']
        ts = data.index[i]
        tm_now = ts.time()

        if tm_now < start_time:
            if verbose:
                print(f"[DIVERGENCE] {ts} skipped (before {start_time.strftime('%H:%M')} cutoff)")
            continue

        stoch_now = data['stk'].iloc[i]
        stoch_known = not pd.isna(stoch_now)
        if not (stoch_known and stoch_now <= stoch_oversold):
            if verbose:
                stoch_str = f"{stoch_now:.1f}" if stoch_known else "n/a"
                print(f"[DIVERGENCE] {ts} skipped (Stoch %K={stoch_str}, not oversold "
                      f"<= {stoch_oversold})")
            continue

        if (i - last_signal_idx) < cooldown_candles:
            if verbose:
                print(f"[DIVERGENCE] {ts} skipped (cooldown)")
            continue

        data.loc[data.index[i], 'st_sig'] = 1
        data.loc[data.index[i], 'agree_count'] = label['agree_count']
        data.loc[data.index[i], 'divergence_indicators'] = label['indicators']
        data.loc[data.index[i], 'reason'] = (
            f"MULTI_INDICATOR_BULLISH_DIVERGENCE | agree={label['agree_count']} "
            f"| indicators=({label['indicators']}) | stoch_k={stoch_now:.1f}"
        )[:220]

        last_signal_idx = i
        fired += 1

        if verbose:
            print(f"[DIVERGENCE] {ts} >>> st_sig=1 | agree={label['agree_count']} "
                  f"| indicators=({label['indicators']}) | stoch_k={stoch_now:.1f}")

    if verbose:
        print(f"\n📊 SIGNALS FIRED: {fired} / {len(pos_label_history)} candidate labels\n")
        if fired:
            fired_rows = data[data['st_sig'] == 1]
            print(f"{'DateTime':<26} {'Close':<10} {'Agree':<7} {'Indicators'}")
            print("-" * 100)
            for ts, row in fired_rows.tail(25).iterrows():
                print(f"{str(ts):<26} {row['Close']:<10.2f} {row['agree_count']:<7} "
                      f"{row['divergence_indicators']}")
            print(f"\n💎 AVG AGREE COUNT: {fired_rows['agree_count'].mean():.2f}")

    return data






# for my market alerts
def tele_msg(msg):
    # Replace YOUR_BOT_TOKEN with your bot token obtained from BotFather
    bot_token1 = Telegram_token.telegram_token

    # Group Chat ID
    chat_id1 = Telegram_token.chat_id

    # Replace MESSAGE_TEXT with the text of the message you want to send
    message_text = " My Market Alerts Super Trend Strategy "+msg

    # Send the message using the sendMessage method of the Telegram Bot API
    url1 = f'https://api.telegram.org/bot{bot_token1}/sendMessage?chat_id={chat_id1}&text={message_text}'
    response = requests.get(url1)


Long_Trade_File = 'SuperTrend_Long.xlsx'
Short_Trade_File = 'SuperTrend_Short.xlsx'


# Create Excel Sheet only of it Needed or first time
def Long_create_excel_sheet(filename):
    if not os.path.exists(filename):
        columns = ['Symbol', 'Entry Time', 'Buy Price', 'Target Price','Sprice', 'Qty', 'Exit Time', 'Sell Price', 'Points',
                   'Brokerage', 'Profit/Loss', 'Trade Status']
        df = pd.DataFrame(columns=columns)
        df.to_excel(filename, index=False)
        print(f"{filename} created successfully!")
    else:
        print(f"{filename} already exists.")


Long_create_excel_sheet(Long_Trade_File)


# Create Excel Sheet only of it Needed or first time
def Short_create_excel_sheet(filename):
    if not os.path.exists(filename):
        columns = ['Symbol', 'Entry Time', 'Buy Price', 'Target Price','Sprice', 'Qty', 'Exit Time', 'Sell Price', 'Points',
                   'Brokerage', 'Profit/Loss', 'Trade Status']
        df = pd.DataFrame(columns=columns)
        df.to_excel(filename, index=False)
        print(f"{filename} created successfully!")
    else:
        print(f"{filename} already exists.")


Short_create_excel_sheet(Short_Trade_File)


def update_target_price(ticker, new_buy_price, tradefile):
    try:
        # Load the Excel file
        df = pd.read_excel(tradefile)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    # Find the row(s) where Symbol matches and Trade Status is OPEN
    index = df[(df['Symbol'] == ticker) & (df['Trade Status'] == 'OPEN')].index

    if not index.empty:
        idx = index[0]  # Only update the first open trade
        df.at[idx, 'Target_Price'] = new_buy_price

        # Save the updated DataFrame
        df.to_excel(tradefile, index=False)
        print(f"Buy Price updated for {ticker}")
    else:
        print(f"No open trade found for {ticker}")

def update_buy_price(ticker, new_buy_price, tradefile):
    try:
        # Load the Excel file
        df = pd.read_excel(tradefile)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    # Find the row(s) where Symbol matches and Trade Status is OPEN
    index = df[(df['Symbol'] == ticker) & (df['Trade Status'] == 'OPEN')].index

    if not index.empty:
        idx = index[0]  # Only update the first open trade
        df.at[idx, 'Sprice'] = new_buy_price

        # Save the updated DataFrame
        df.to_excel(tradefile, index=False)
        print(f"Buy Price updated for {ticker}")
    else:
        print(f"No open trade found for {ticker}")
# Update Long trade entry
def update_long_trades(ticker, entry_time, BuyPrice, target_price,Sprice ,qty, tradefile):
    workbook = load_workbook(filename=tradefile)
    worksheet = workbook.active

    new_row = [ticker, entry_time, BuyPrice, target_price,Sprice, qty, '', '', '', '', '', 'OPEN']
    worksheet.append(new_row)

    workbook.save(tradefile)


# Update Long trade entry
def update_Short_trades(ticker, entry_time, SellPrice, target_price,Sprice, qty, filename):
    workbook = load_workbook(filename)
    worksheet = workbook.active

    new_row = [ticker, entry_time, '', target_price,Sprice, qty, '', SellPrice, '', '', '', 'OPEN']
    worksheet.append(new_row)

    workbook.save(filename)


def all_trade_files():
    files = [Long_Trade_File, Short_Trade_File]
    merged_df = pd.concat([pd.read_excel(file) for file in files], ignore_index=True)
    merged_df.to_excel('All_Trades.xlsx', index=False)
    return 'All_Trades.xlsx saved successfully.'


# Call the function
all_trade_files()


# data = get_cash_market_data('INFY', '5m')
# super_trend(data)


# Symbol = 'INFY'
# Entry_Time = '2024-12-30 09:35:00'
# Buy_Price = 350
# Target_Price = 377
# Qty = 600
# Exit_Time = '2025-01-29 10:40:00'
# Sell_Price = 378
# Points = Sell_Price - Buy_Price
# Brokerage = (Buy_Price*Qty)+(Sell_Price*Qty) * 0.00015
# Profit_Loss = (Points * Qty) - Brokerage
# Trade_Status = 'Target Hit'

# update_long_trades(Symbol,Entry_Time,Buy_Price,Target_Price,Qty,Long_Trade_File)

# Function to close long trade
def close_long_trade(ticker, exit_time, sell_price, points, brokerage, profit_loss, trade_status, tradefile):
    df = pd.read_excel(tradefile)
    index = df[(df['Symbol'] == ticker) & (df['Trade Status'] == 'OPEN')].index

    if not index.empty:
        idx = index[0]
        df.at[idx, 'Exit Time'] = exit_time
        df.at[idx, 'Sell_Price'] = sell_price
        df.at[idx, 'Points'] = points
        df.at[idx, 'Brokerage'] = brokerage
        df.at[idx, 'Profit/Loss'] = profit_loss
        df.at[idx, 'Trade Status'] = trade_status

        df.to_excel(tradefile, index=False)
        print(f"Trade closed for {ticker}")
    else:
        print(f"No open trade found for {ticker}")


# Function to close long trade
def close_short_trade(ticker, exit_time, buy_price, points, brokerage, profit_loss, trade_status, tradefile):
    df = pd.read_excel(tradefile)
    index = df[(df['Symbol'] == ticker) & (df['Trade Status'] == 'OPEN')].index

    if not index.empty:
        idx = index[0]
        df.at[idx, 'Exit Time'] = exit_time
        df.at[idx, 'Buy_Price'] = buy_price
        df.at[idx, 'Points'] = points
        df.at[idx, 'Brokerage'] = brokerage
        df.at[idx, 'Profit/Loss'] = profit_loss
        df.at[idx, 'Trade Status'] = trade_status

        df.to_excel(tradefile, index=False)
        print(f"Trade closed for {ticker}")
    else:
        print(f"No open trade found for {ticker}")


# Define the required times
required_times =  [
    (9, 15), (9, 18), (9, 21), (9, 24), (9, 27), (9, 30), (9, 33), (9, 36), (9, 39), (9, 42), (9, 45), (9, 48), (9, 51), (9, 54), (9, 57),
    (10, 0), (10, 3), (10, 6), (10, 9), (10, 12), (10, 15), (10, 18), (10, 21), (10, 24), (10, 27), (10, 30), (10, 33), (10, 36), (10, 39), (10, 42), (10, 45), (10, 48), (10, 51), (10, 54), (10, 57),
    (11, 0), (11, 3), (11, 6), (11, 9), (11, 12), (11, 15), (11, 18), (11, 21), (11, 24), (11, 27), (11, 30), (11, 33), (11, 36), (11, 39), (11, 42), (11, 45), (11, 48), (11, 51), (11, 54), (11, 57),
    (12, 0), (12, 3), (12, 6), (12, 9), (12, 12), (12, 15), (12, 18), (12, 21), (12, 24), (12, 27), (12, 30), (12, 33), (12, 36), (12, 39), (12, 42), (12, 45), (12, 48), (12, 51), (12, 54), (12, 57),
    (13, 0), (13, 3), (13, 6), (13, 9), (13, 12), (13, 15), (13, 18), (13, 21), (13, 24), (13, 27), (13, 30), (13, 33), (13, 36), (13, 39), (13, 42), (13, 45), (13, 48), (13, 51), (13, 54), (13, 57),
    (14, 0), (14, 3), (14, 6), (14, 9), (14, 12), (14, 15), (14, 18), (14, 21), (14, 24), (14, 27), (14, 30), (14, 33), (14, 36), (14, 39), (14, 42), (14, 45), (14, 48), (14, 51), (14, 54), (14, 57),
    (15, 0), (15, 3), (15, 6), (15, 9), (15, 12), (15, 15), (15, 18), (15, 21), (15, 24), (15, 27), (15, 30)
]


# Define a function to check if the current time matches any of the required times
def is_required_time():
    current_time = dt.datetime.now(pytz.timezone('Asia/Kolkata')).time()
    return any(current_time.hour == hour and current_time.minute == minute for hour, minute in required_times)


# Initialize spot prices dictionary
spot_prices1 = {ticker: None for ticker in Tickers}

# Get instrument codes for tickers
ticker_codes = {ticker: str(instrument_df[instrument_df['Name'] == ticker]['ScripCode'].values[0]) for ticker in
                Tickers}

# Getting the List for Streaming Data
req_list = []
for s in Tickers:
    code = str(scripcode_lookup(instrument=instrument_df, symbol=s))
    req = {"Exch": "N", "ExchType": "D", "ScripCode": code}

    req_list.append(req)


# Define the callback function for incoming data
def on_message(ws, message):
    global spot_prices1
    data = json.loads(message)
    #print(data)
    if data:
        ticker_symbol = instrument_df[instrument_df['ScripCode'] == data[0]['Token']]['Name'].iloc[0]
        last_rate = data[0]['LastRate']
        spot_prices1[ticker_symbol] = last_rate

    # Subscribe to real-time data feed in a separate thread


# req_list = [{"Exch": "N", "ExchType": "C", "ScripCode": code} for code in ticker_codes.values()]
req_data = client.Request_Feed('mf', 's', req_list)
client.connect(req_data)


# function for Subscribing Data
def subscribe_data():
    client.receive_data(on_message)

import requests

def send_to_ui(symbol: str, price: float):
    try:
        # First, try to update the ticker
        res = requests.post("https://algotrading-ufvn.onrender.com/update-ticker", json={
            "symbol": symbol,
            "price": price
        })

        # If symbol is not found (not added yet), auto-add and retry
        if res.status_code == 404:
            print(f"⚠️ {symbol} not found in ticker list. Adding it now...")
            add_res = requests.post("http://localhost:8000/add-tickers", json={
                "tickers": [symbol]
            })
            if add_res.status_code == 200:
                print(f"✅ {symbol} added. Retrying update...")
                # Retry update
                retry_res = requests.post("http://localhost:8000/update-ticker", json={
                    "symbol": symbol,
                    "price": price
                })
                if retry_res.status_code == 200:
                    print(f"📈 Pushed to UI: {symbol} → ₹{price}")
                else:
                    print(f"❌ Failed to push after retry: {retry_res.status_code}")
            else:
                print(f"❌ Failed to auto-add {symbol}: {add_res.status_code}")

        elif res.status_code == 200:
            print(f"📈 Pushed to UI: {symbol} → ₹{price}")
        else:
            print(f"❌ Failed to push: {res.status_code}")
    except Exception as e:
        print(f"🔥 Error pushing to UI: {e}")
# Algo Start Here
start = dt.datetime.now(pytz.timezone('Asia/Kolkata'))
closetime = start.replace(hour=START_TIME[0], minute=START_TIME[1], second=START_TIME[2])
interval = (closetime-start).total_seconds()
if interval > 0:
    print('Algo will Run at ', START_TIME[0], ':', START_TIME[1], ':', START_TIME[2], ' Remaining Time Left = ',
          interval, ' sec')
    time.sleep(interval)
    print('Algo Starting Now!!!')

data_list = {}

for h in Tickers:

    data_fut = get_cash_market_data(h, '3m')
    data_fut.drop(data_fut.tail(1).index, inplace=True)
    df = get_cash_market_data_3('3m')
    data_fut =super_trend(h,data_fut)

    data_list[h] = data_fut

super_Trend_Long = pd.read_excel(Long_Trade_File)
Long_Open_Position = super_Trend_Long[(super_Trend_Long['Trade Status'] == 'OPEN')]

super_Trend_Short = pd.read_excel(Short_Trade_File)
Short_Open_Position = super_Trend_Short[(super_Trend_Short['Trade Status'] == 'OPEN')]

# Adding Multi Thread Process to run Streaming data and Main algo at the same time
streaming_thread = threading.Thread(target=subscribe_data)
streaming_thread.daemon = True
streaming_thread.start()

time.sleep(5)

endTime = dt.datetime.now(pytz.timezone('Asia/Kolkata')).replace(hour=EXIT_TIME[0], minute=EXIT_TIME[1],
                                                                 second=EXIT_TIME[2])
# ── track processed candle minutes to avoid double-fire ──────
last_processed_minute = {}  # {ticker: (hour, minute)}



while dt.datetime.now(pytz.timezone('Asia/Kolkata')) < endTime:

    try:

        for i in Tickers:

            now_ist = dt.datetime.now(pytz.timezone('Asia/Kolkata'))
            current_minute = (now_ist.hour, now_ist.minute)

            if spot_prices1[i] is None:
                time.sleep(0.5)
                continue

            send_to_ui(i, spot_prices1[i])
            time.sleep(0.5)

            # ════════════════════════════════════════════════════
            # BLOCK A — SIGNAL + ENTRY + REVERSAL (CE <-> PE)
            # Runs only once per 5-min candle per ticker
            # ════════════════════════════════════════════════════
            if is_required_time() and last_processed_minute.get(i) != current_minute:

                last_processed_minute[i] = current_minute  # mark this candle done

                try:
                    data_fut = get_cash_market_data(i, '3m')
                    print("###################################################################")
                    print('Spot prices of', i, ' ', spot_prices1[i])
                    print(now_ist.strftime("%d-%b-%Y %I:%M:%S%p"))
                    print("###################################################################")
                except Exception as e:
                    print(f"⚠️ Skipping {i}: {e}")
                    continue

                data_fut.drop(data_fut.tail(1).index, inplace=True)
                df = get_cash_market_data_3('3m')
                data_fut = super_trend(i, data_fut, 4, df)
                data_list[i] = data_fut

                super_Trend_Long   = pd.read_excel(Long_Trade_File)
                Long_Open_Position = super_Trend_Long[super_Trend_Long['Trade Status'] == 'OPEN']
                super_Trend_Short   = pd.read_excel(Short_Trade_File)
                Short_Open_Position = super_Trend_Short[super_Trend_Short['Trade Status'] == 'OPEN']

                # ── Pull the signal from only the latest completed candle ──
                recent = data_list[i].iloc[[-1]]
                signal_rows = recent[recent["st_sig"] == 1]

                if not signal_rows.empty:

                    latest_signal_row = signal_rows.iloc[-1]
                    Option_Type = latest_signal_row["Option_Type"]  # 'CE' or 'PE'

                    is_ce_open = i in Long_Open_Position['Symbol'].values
                    is_pe_open = i in Short_Open_Position['Symbol'].values

                    current_price  = float(spot_prices1[i])
                    Trade_quantity = 65
                    entry_time     = dt.datetime.now(UTC).strftime("%d-%b-%Y %I:%M%p")

                    # ────────────────────────────────────────────
                    # SAME-TYPE signal while that type already open → block, do nothing else
                    # ────────────────────────────────────────────
                    if (Option_Type == 'CE' and is_ce_open) or (Option_Type == 'PE' and is_pe_open):
                        print(f"{i} already in {Option_Type} Open Position. Maximum Position Reached. No New Position.")
                        continue

                    # ────────────────────────────────────────────
                    # OPPOSITE-TYPE signal → reversal (exit current leg first)
                    # ────────────────────────────────────────────
                    if Option_Type == 'PE' and is_ce_open:
                        print(f"🔄 Reversal: {i} CE → PE. Exiting CE first.")
                        trade_row           = Long_Open_Position[Long_Open_Position['Symbol'] == i]
                        BuyPrice_exit       = float(trade_row['Buy Price'].values[0])
                        Trade_quantity_exit = int(trade_row['Qty'].values[0])
                        Exit_Time           = dt.datetime.now(UTC).strftime("%d-%b-%Y %I:%M%p")
                        Sell_Price          = current_price
                        Points              = Sell_Price - BuyPrice_exit
                        Brokerage           = ((BuyPrice_exit * Trade_quantity_exit) + (Sell_Price * Trade_quantity_exit)) * 0.00015
                        Profit_Loss         = (Points * Trade_quantity_exit) - Brokerage

                        close_long_trade(i, Exit_Time, Sell_Price, Points, Brokerage, Profit_Loss, "Reversal Exit", Long_Trade_File)
                        tele_msg(f"🔄 Reversal Exit (CE→PE): {i} | Exit: {Sell_Price} | P/L: {round(Profit_Loss,2)}")

                        super_Trend_Long   = pd.read_excel(Long_Trade_File)
                        Long_Open_Position = super_Trend_Long[super_Trend_Long['Trade Status'] == 'OPEN']
                        is_ce_open = False

                    elif Option_Type == 'CE' and is_pe_open:
                        print(f"🔄 Reversal: {i} PE → CE. Exiting PE first.")
                        trade_row           = Short_Open_Position[Short_Open_Position['Symbol'] == i]
                        BuyPrice_exit       = float(trade_row['Buy Price'].values[0])
                        Trade_quantity_exit = int(trade_row['Qty'].values[0])
                        Exit_Time           = dt.datetime.now(UTC).strftime("%d-%b-%Y %I:%M%p")
                        Sell_Price          = current_price
                        Points              = Sell_Price - BuyPrice_exit
                        Brokerage           = ((BuyPrice_exit * Trade_quantity_exit) + (Sell_Price * Trade_quantity_exit)) * 0.00015
                        Profit_Loss         = (Points * Trade_quantity_exit) - Brokerage

                        close_short_trade(i, Exit_Time, Sell_Price, Points, Brokerage, Profit_Loss, "Reversal Exit", Short_Trade_File)
                        tele_msg(f"🔄 Reversal Exit (PE→CE): {i} | Exit: {Sell_Price} | P/L: {round(Profit_Loss,2)}")

                        super_Trend_Short   = pd.read_excel(Short_Trade_File)
                        Short_Open_Position = super_Trend_Short[super_Trend_Short['Trade Status'] == 'OPEN']
                        is_pe_open = False

                    # ────────────────────────────────────────────
                    # FRESH ENTRY (new ticker, or reversal just cleared the opposite leg)
                    # ────────────────────────────────────────────
                    all_trade_files()
                    open_trades_df    = pd.read_excel('All_Trades.xlsx')
                    open_trade_count  = len(open_trades_df[open_trades_df['Trade Status'] == 'OPEN'])

                    if open_trade_count >= Max_Position:
                        print("Maximum Position Reached. No New Position.")
                        continue

                    Target_Price = current_price + Take_Profit
                    Sprice       = current_price - 10
                    BuyPrice     = current_price

                    if Option_Type == 'CE':
                        update_long_trades(i, entry_time, BuyPrice, Target_Price, Sprice, Trade_quantity, Long_Trade_File)
                        tele_msg(
                            f"CE Entry: {i} | Qty: {Trade_quantity} | "
                            f"Buy: {BuyPrice} | Target: {Target_Price} | SL: {Sprice}"
                        )
                        super_Trend_Long   = pd.read_excel(Long_Trade_File)
                        Long_Open_Position = super_Trend_Long[super_Trend_Long['Trade Status'] == 'OPEN']

                    elif Option_Type == 'PE':
                        update_Short_trades(i, entry_time, BuyPrice, Target_Price, Sprice, Trade_quantity, Short_Trade_File)
                        tele_msg(
                            f"PE Entry: {i} | Qty: {Trade_quantity} | "
                            f"Buy: {BuyPrice} | Target: {Target_Price} | SL: {Sprice}"
                        )
                        super_Trend_Short   = pd.read_excel(Short_Trade_File)
                        Short_Open_Position = super_Trend_Short[super_Trend_Short['Trade Status'] == 'OPEN']

            # ════════════════════════════════════════════════════
            # BLOCK B — SL / TARGET / TRAILING / TIMEOUT  (CE leg)
            # Runs EVERY loop iteration (~every 0.5s)
            # ════════════════════════════════════════════════════

            super_Trend_Long   = pd.read_excel(Long_Trade_File)
            Long_Open_Position = super_Trend_Long[super_Trend_Long['Trade Status'] == 'OPEN']

            if i in Long_Open_Position['Symbol'].values:

                trade_row      = Long_Open_Position[Long_Open_Position['Symbol'] == i]
                BuyPrice       = float(trade_row['Buy Price'].values[0])
                Target_Price   = float(trade_row['Target Price'].values[0])
                S_Price        = float(trade_row['Sprice'].values[0])
                Trade_quantity = int(trade_row['Qty'].values[0])
                current_price  = float(spot_prices1[i])

                profit_from_entry = current_price - BuyPrice

                print(f"  📊 CE {i} | Buy:{BuyPrice} | Now:{current_price} | "
                      f"Target:{Target_Price} | SL:{S_Price} | PnL:{round(profit_from_entry,2)}")

                # ── B1: TRAILING SL + STAGED TARGET ───────────────
                Target_Step        = 10
                SL_Step             = 5
                Max_Target_Offset  = 50

                if profit_from_entry >= SL_Step:

                    target_steps = int(profit_from_entry // Target_Step)
                    new_target   = BuyPrice + min((target_steps + 1) * Target_Step, Max_Target_Offset)
                    new_target   = max(new_target, Target_Price)

                    sl_steps = int(profit_from_entry // SL_Step)
                    new_sl   = BuyPrice + ((sl_steps - 2) * SL_Step)
                    new_sl   = max(new_sl, S_Price)

                    if new_sl > S_Price or new_target > Target_Price:
                        print(f"  📈 TRAILING SL: {S_Price} → {new_sl}  |  TARGET: {Target_Price} → {new_target}  [CE {i}]")
                        tele_msg(f"📈 CE Trail Update: {i} | SL: {S_Price}→{new_sl} | Target: {Target_Price}→{new_target}")

                        update_buy_price(i, new_sl, Long_Trade_File)
                        update_target_price(i, new_target, Long_Trade_File)

                        super_Trend_Long   = pd.read_excel(Long_Trade_File)
                        Long_Open_Position = super_Trend_Long[super_Trend_Long['Trade Status'] == 'OPEN']
                        trade_row          = Long_Open_Position[Long_Open_Position['Symbol'] == i]
                        S_Price            = float(trade_row['Sprice'].values[0])
                        Target_Price       = float(trade_row['Target Price'].values[0])

                # ── B2: HARD SL HIT ──────────────────────────────
                if current_price <= S_Price:
                    print(f"  🔴 SL HIT → CE {i} | Price:{current_price} | SL:{S_Price}")
                    Exit_Time    = dt.datetime.now(UTC).strftime("%d-%b-%Y %I:%M%p")
                    Sell_Price   = current_price
                    Points       = Sell_Price - BuyPrice
                    Brokerage    = ((BuyPrice * Trade_quantity) + (Sell_Price * Trade_quantity)) * 0.00015
                    Profit_Loss  = (Points * Trade_quantity) - Brokerage
                    Trade_Status = "SL Hit"

                    close_long_trade(i, Exit_Time, Sell_Price, Points, Brokerage, Profit_Loss, Trade_Status, Long_Trade_File)
                    tele_msg(f"🔴 CE SL Hit: {i} | Exit: {Sell_Price} | P/L: {Profit_Loss}")
                    continue

                # ── B3: TARGET HIT ────────────────────────────────
                if current_price >= Target_Price:
                    print(f"  🎯 TARGET HIT → CE {i} | Price:{current_price} | Target:{Target_Price}")
                    Exit_Time    = dt.datetime.now(UTC).strftime("%d-%b-%Y %I:%M%p")
                    Sell_Price   = current_price
                    Points       = Sell_Price - BuyPrice
                    Brokerage    = ((BuyPrice * Trade_quantity) + (Sell_Price * Trade_quantity)) * 0.00015
                    Profit_Loss  = (Points * Trade_quantity) - Brokerage
                    Trade_Status = "Target Hit"

                    close_long_trade(i, Exit_Time, Sell_Price, Points, Brokerage, Profit_Loss, Trade_Status, Long_Trade_File)
                    tele_msg(f"🎯 CE Target Hit: {i} | Exit: {Sell_Price} | P/L: {Profit_Loss}")
                    continue

                # ── B4: EXIT TIME OUT ─────────────────────────────
                exit_cutoff = now_ist.replace(
                    hour=EXIT_TIME[0],
                    minute=max(EXIT_TIME[1] - 5, 0),
                    second=EXIT_TIME[2]
                )
                if now_ist >= exit_cutoff:
                    print(f"  ⏰ TIME OUT → CE {i}")
                    Exit_Time    = dt.datetime.now(UTC).strftime("%d-%b-%Y %I:%M%p")
                    Sell_Price   = current_price
                    Points       = Sell_Price - BuyPrice
                    Brokerage    = ((BuyPrice * Trade_quantity) + (Sell_Price * Trade_quantity)) * 0.00015
                    Profit_Loss  = (Points * Trade_quantity) - Brokerage
                    Trade_Status = "Exit Time Out"

                    close_long_trade(i, Exit_Time, Sell_Price, Points, Brokerage, Profit_Loss, Trade_Status, Long_Trade_File)
                    tele_msg(f"⏰ CE Time Out: {i} | Exit: {Sell_Price} | P/L: {Profit_Loss}")
                    continue

            # ════════════════════════════════════════════════════
            # BLOCK C — SL / TARGET / TRAILING / TIMEOUT  (PE leg)
            # Mirrors Block B exactly, reading/writing Short_Trade_File
            # ════════════════════════════════════════════════════

            super_Trend_Short   = pd.read_excel(Short_Trade_File)
            Short_Open_Position = super_Trend_Short[super_Trend_Short['Trade Status'] == 'OPEN']

            if i not in Short_Open_Position['Symbol'].values:
                continue  # nothing open on either leg, move to next ticker

            trade_row      = Short_Open_Position[Short_Open_Position['Symbol'] == i]
            BuyPrice       = float(trade_row['Buy Price'].values[0])
            Target_Price   = float(trade_row['Target Price'].values[0])
            S_Price        = float(trade_row['Sprice'].values[0])
            Trade_quantity = int(trade_row['Qty'].values[0])
            current_price  = float(spot_prices1[i])

            profit_from_entry = current_price - BuyPrice

            print(f"  📊 PE {i} | Buy:{BuyPrice} | Now:{current_price} | "
                  f"Target:{Target_Price} | SL:{S_Price} | PnL:{round(profit_from_entry,2)}")

            # ── C1: TRAILING SL + STAGED TARGET ───────────────
            Target_Step        = 10
            SL_Step             = 5
            Max_Target_Offset  = 50

            if profit_from_entry >= SL_Step:

                target_steps = int(profit_from_entry // Target_Step)
                new_target   = BuyPrice + min((target_steps + 1) * Target_Step, Max_Target_Offset)
                new_target   = max(new_target, Target_Price)

                sl_steps = int(profit_from_entry // SL_Step)
                new_sl   = BuyPrice + ((sl_steps - 2) * SL_Step)
                new_sl   = max(new_sl, S_Price)

                if new_sl > S_Price or new_target > Target_Price:
                    print(f"  📈 TRAILING SL: {S_Price} → {new_sl}  |  TARGET: {Target_Price} → {new_target}  [PE {i}]")
                    tele_msg(f"📈 PE Trail Update: {i} | SL: {S_Price}→{new_sl} | Target: {Target_Price}→{new_target}")

                    update_buy_price(i, new_sl, Short_Trade_File)
                    update_target_price(i, new_target, Short_Trade_File)

                    super_Trend_Short   = pd.read_excel(Short_Trade_File)
                    Short_Open_Position = super_Trend_Short[super_Trend_Short['Trade Status'] == 'OPEN']
                    trade_row           = Short_Open_Position[Short_Open_Position['Symbol'] == i]
                    S_Price             = float(trade_row['Sprice'].values[0])
                    Target_Price        = float(trade_row['Target Price'].values[0])

            # ── C2: HARD SL HIT ──────────────────────────────
            if current_price <= S_Price:
                print(f"  🔴 SL HIT → PE {i} | Price:{current_price} | SL:{S_Price}")
                Exit_Time    = dt.datetime.now(UTC).strftime("%d-%b-%Y %I:%M%p")
                Sell_Price   = current_price
                Points       = Sell_Price - BuyPrice
                Brokerage    = ((BuyPrice * Trade_quantity) + (Sell_Price * Trade_quantity)) * 0.00015
                Profit_Loss  = (Points * Trade_quantity) - Brokerage
                Trade_Status = "SL Hit"

                close_short_trade(i, Exit_Time, Sell_Price, Points, Brokerage, Profit_Loss, Trade_Status, Short_Trade_File)
                tele_msg(f"🔴 PE SL Hit: {i} | Exit: {Sell_Price} | P/L: {Profit_Loss}")
                continue

            # ── C3: TARGET HIT ────────────────────────────────
            if current_price >= Target_Price:
                print(f"  🎯 TARGET HIT → PE {i} | Price:{current_price} | Target:{Target_Price}")
                Exit_Time    = dt.datetime.now(UTC).strftime("%d-%b-%Y %I:%M%p")
                Sell_Price   = current_price
                Points       = Sell_Price - BuyPrice
                Brokerage    = ((BuyPrice * Trade_quantity) + (Sell_Price * Trade_quantity)) * 0.00015
                Profit_Loss  = (Points * Trade_quantity) - Brokerage
                Trade_Status = "Target Hit"

                close_short_trade(i, Exit_Time, Sell_Price, Points, Brokerage, Profit_Loss, Trade_Status, Short_Trade_File)
                tele_msg(f"🎯 PE Target Hit: {i} | Exit: {Sell_Price} | P/L: {Profit_Loss}")
                continue

            # ── C4: EXIT TIME OUT ─────────────────────────────
            exit_cutoff = now_ist.replace(
                hour=EXIT_TIME[0],
                minute=max(EXIT_TIME[1] - 5, 0),
                second=EXIT_TIME[2]
            )
            if now_ist >= exit_cutoff:
                print(f"  ⏰ TIME OUT → PE {i}")
                Exit_Time    = dt.datetime.now(UTC).strftime("%d-%b-%Y %I:%M%p")
                Sell_Price   = current_price
                Points       = Sell_Price - BuyPrice
                Brokerage    = ((BuyPrice * Trade_quantity) + (Sell_Price * Trade_quantity)) * 0.00015
                Profit_Loss  = (Points * Trade_quantity) - Brokerage
                Trade_Status = "Exit Time Out"

                close_short_trade(i, Exit_Time, Sell_Price, Points, Brokerage, Profit_Loss, Trade_Status, Short_Trade_File)
                tele_msg(f"⏰ PE Time Out: {i} | Exit: {Sell_Price} | P/L: {Profit_Loss}")
                continue

    except Exception as e:
        ct = dt.datetime.now().strftime("%d-%b-%Y %I:%M%p")
        error_message = f"{ct} - An error occurred: {e}"
        print("Error:", e)
        print("Oops!", e.__class__, "occurred.")
        tele_msg(error_message)
        with open("error_log.txt", "a") as error_log_file:
            error_log_file.write(error_message + "\n")
        # ✅ NO raise — bot stays alive on any error
        time.sleep(2)
